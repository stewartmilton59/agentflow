import csv
import openpyxl
from decimal import Decimal
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, F, Sum
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.template.loader import get_template
import uuid
from accounts.decorators import role_required
from sales.models import SaleItem, Sale
from .models import (
    Category, Product, StockMovement, StockAlert, InventoryAdjustment
)
from .forms import (
    CategoryForm, ProductForm, StockAdjustmentForm, ProductSearchForm
)
from core.models import ActivityLog, Company, PaymentMethod


# ==============================================================================
# PRODUCT VIEWS
# ==============================================================================

@login_required
def product_list_view(request):
    """List all products"""
    products = Product.objects.filter(is_active=True).select_related('category').only(
        'id', 'name', 'sku', 'generic_name', 'product_type', 'purchase_price',
        'selling_price', 'current_stock', 'reorder_level', 'category',
        'expiry_date', 'batch_number', 'pack_size', 'is_active', 'image',
        'prescription_required'
    )

    search = request.GET.get('search', '')
    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(generic_name__icontains=search)
        )

    product_type = request.GET.get('product_type')
    if product_type:
        products = products.filter(product_type=product_type)

    low_stock = request.GET.get('low_stock')
    if low_stock:
        products = products.filter(current_stock__lte=F('reorder_level'), current_stock__gt=0)

    out_of_stock = request.GET.get('out_of_stock')
    if out_of_stock:
        products = products.filter(current_stock=0)

    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate dates for expiry warnings
    from datetime import date, timedelta
    today = date.today()
    expiry_warning_date = today + timedelta(days=30)  # Warning for products expiring in 30 days

    context = {
        'products': page_obj,
        'search': search,
        'product_type': product_type,
        'product_types': Product.PRODUCT_TYPES,
        'low_stock_filter': low_stock,
        'out_of_stock_filter': out_of_stock,
        'today': today,
        'expiry_warning_date': expiry_warning_date,
    }
    return render(request, 'inventory/product_list.html', context)


@login_required
def product_create_view(request):
    """Create new product"""
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                product = form.save(commit=False)
                product.current_stock = 0
                product.created_by = request.user

                # Set default values if not provided
                if not product.unit:
                    product.unit = 'piece'
                if not product.reorder_level:
                    product.reorder_level = 10
                if not product.minimum_stock:
                    product.minimum_stock = 5
                if not product.maximum_stock:
                    product.maximum_stock = 500

                product.save()

                ActivityLog.objects.create(
                    user=request.user,
                    action='create',
                    model_name='Product',
                    object_id=str(product.id),
                    object_repr=product.name,
                    ip_address=request.META.get('REMOTE_ADDR')
                )

                messages.success(request, f'Product "{product.name}" created successfully!')
                return redirect('inventory:product_detail', pk=product.id)

            except Exception as e:
                messages.error(request, f'Error saving product: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ProductForm()

    context = {
        'form': form,
        'title': 'Add New Product',
    }
    return render(request, 'inventory/product_form.html', context)


@login_required
def product_detail_view(request, pk):
    """View product details"""
    product = get_object_or_404(Product, id=pk)

    # Get recent stock movements for this product
    movements = StockMovement.objects.filter(product=product).select_related('created_by').only(
        'id', 'movement_type', 'quantity', 'previous_quantity', 'new_quantity',
        'unit_price', 'total_amount', 'reference_type', 'reference_id', 'notes',
        'created_by', 'created_at'
    ).order_by('-created_at')[:10]

    context = {
        'product': product,
        'movements': movements,
    }
    return render(request, 'inventory/product_detail.html', context)


@login_required
def product_update_view(request, pk):
    """Update product"""
    product = get_object_or_404(Product, id=pk)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            try:
                updated_product = form.save(commit=False)
                updated_product.save()

                ActivityLog.objects.create(
                    user=request.user,
                    action='update',
                    model_name='Product',
                    object_id=str(product.id),
                    object_repr=product.name,
                    ip_address=request.META.get('REMOTE_ADDR')
                )

                messages.success(request, f'Product "{updated_product.name}" updated successfully!')
                return redirect('inventory:product_detail', pk=product.id)
            except Exception as e:
                messages.error(request, f'Error updating product: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ProductForm(instance=product)

    context = {
        'form': form,
        'product': product,
        'title': 'Edit Product',
    }
    return render(request, 'inventory/product_form.html', context)


@login_required
def product_delete_view(request, pk):
    """Delete product"""
    product = get_object_or_404(Product, id=pk)

    if request.method == 'POST':
        product_name = product.name
        product.delete()

        ActivityLog.objects.create(
            user=request.user,
            action='delete',
            model_name='Product',
            object_id=str(pk),
            object_repr=product_name,
            ip_address=request.META.get('REMOTE_ADDR')
        )

        messages.success(request, f'Product "{product_name}" deleted successfully!')
        return redirect('inventory:product_list')

    return render(request, 'inventory/product_confirm_delete.html', {'product': product})


# ==============================================================================
# STOCK MANAGEMENT VIEWS
# ==============================================================================

@login_required
def stock_list_view(request):
    """Display current stock levels for all products"""
    products = Product.objects.filter(is_active=True).select_related('category').only(
        'id', 'name', 'sku', 'current_stock', 'reorder_level', 'selling_price',
        'purchase_price', 'category', 'expiry_date', 'batch_number', 'pack_size',
        'is_active', 'image'
    )

    search = request.GET.get('search', '')
    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(generic_name__icontains=search)
        )

    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)

    low_stock = request.GET.get('low_stock')
    if low_stock:
        products = products.filter(current_stock__lte=F('reorder_level'))

    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    categories = Category.objects.filter(is_active=True)

    context = {
        'products': page_obj,
        'search': search,
        'selected_category': category_id,
        'low_stock_filter': low_stock,
        'categories': categories,
    }

    return render(request, 'inventory/stock_list.html', context)

# ================================== stock_movements_fuctions ==================
@login_required
def stock_movement_view(request):
    """View stock movements with filtering by type, product (Name/SKU/UUID), and date range."""
    movements = (
        StockMovement.objects.select_related("product", "created_by")
        .all()
        .order_by("-created_at")
    )

    # 1. Filter by Movement Type (purchase, sale, adjustment, etc.)
    movement_type = request.GET.get("type")
    if movement_type:
        movements = movements.filter(movement_type=movement_type)

    # 2. Filter by Product Name, SKU, or UUID
    product_query = request.GET.get("product")
    if product_query:
        try:
            # Check if input is a valid UUID
            valid_uuid = uuid.UUID(product_query)
            movements = movements.filter(product_id=valid_uuid)
        except ValueError:
            # Not a UUID -> Search case-insensitively by Product Name or SKU
            movements = movements.filter(
                Q(product__name__icontains=product_query)
                | Q(product__sku__icontains=product_query)
            )

    # 3. Filter by Date Range
    date_from = request.GET.get("date_from")
    if date_from:
        movements = movements.filter(created_at__date__gte=date_from)

    date_to = request.GET.get("date_to")
    if date_to:
        movements = movements.filter(created_at__date__lte=date_to)

    # 4. Pagination
    paginator = Paginator(movements, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "movements": page_obj,
        "selected_type": movement_type or "",
        "product_query": product_query or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
    }

    return render(request, "inventory/stock_movement.html", context)


# ================= stock_movement_print_view ===================================


@login_required
def stock_movement_print_view(request):
    """Print stock movements in invoice format (browser print)."""
    movements = (
        StockMovement.objects.select_related("product", "created_by")
        .all()
        .order_by("-created_at")
    )

    movement_type = request.GET.get("type")
    if movement_type:
        movements = movements.filter(movement_type=movement_type)

    product_query = request.GET.get("product")
    if product_query:
        try:
            valid_uuid = uuid.UUID(product_query)
            movements = movements.filter(product_id=valid_uuid)
        except ValueError:
            movements = movements.filter(
                Q(product__name__icontains=product_query)
                | Q(product__sku__icontains=product_query)
            )

    date_from = request.GET.get("date_from")
    if date_from:
        movements = movements.filter(created_at__date__gte=date_from)

    date_to = request.GET.get("date_to")
    if date_to:
        movements = movements.filter(created_at__date__lte=date_to)

    company = Company.objects.first()

    pharmacy_info = {
        'name': company.name if company else 'Robby One Pharmacy',
        'p_o_box': getattr(company, 'p_o_box', '') if company else '',
        'address': (f"P.O.BOX {company.p_o_box} {company.location}" if company and company.p_o_box and company.location else
                    getattr(company, 'address', '') if company else ''),
        'phone': getattr(company, 'phone', '') if company else '',
        'email': getattr(company, 'email', '') if company else '',
        'tin': getattr(company, 'tax_id', '') if company else '',
        'location': getattr(company, 'location', '') if company else '',
    }

    bank_accounts = list(
        company.payment_methods.filter(is_active=True).values('bank_name', 'account_name', 'account_number')
        if company else []
    )

    totals = movements.aggregate(
        total_movements=Count('id'),
        total_quantity=Sum('quantity'),
        type_totals_amount=Sum('total_amount'),
    )
    total_movements = totals['total_movements'] or 0
    total_quantity = totals['total_quantity'] or 0
    type_totals_amount = totals['type_totals_amount'] or 0

    type_breakdown = []
    type_aggs = movements.values('movement_type').annotate(
        count=Count('id'),
        qty=Sum('quantity'),
        amt=Sum('total_amount'),
    )
    type_map = dict(StockMovement.MOVEMENT_TYPES)
    for row in type_aggs:
        type_breakdown.append({
            'type': type_map.get(row['movement_type'], row['movement_type']),
            'count': row['count'],
            'total_qty': row['qty'] or 0,
            'total_amt': row['amt'] or 0,
        })

    context = {
        'movements': movements,
        'company': company,
        'pharmacy_info': pharmacy_info,
        'bank_accounts': bank_accounts,
        'report_date': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'generated_by': request.user.get_full_name() or request.user.username,
        'total_movements': total_movements,
        'total_quantity': total_quantity,
        'type_totals_amount': type_totals_amount,
        'type_breakdown': type_breakdown,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'selected_type': dict(StockMovement.MOVEMENT_TYPES).get(movement_type, '') if movement_type else '',
        'product_query': product_query or '',
    }
    return render(request, "inventory/stock_movement_print.html", context)


# ================= stock_alerts_fuction =======================================

@login_required
def stock_alerts_view(request):
    """Auto-generate and view active stock alerts"""
    if request.method == 'POST' and request.POST.get('resolve'):
        alert_id = request.POST.get('alert_id')
        if alert_id:
            StockAlert.objects.filter(id=alert_id, status='active').update(
                status='resolved',
                resolved_at=timezone.now(),
                resolved_by=request.user,
            )
            messages.success(request, 'Alert resolved.')
            return redirect('inventory:stock_alerts')

    now = timezone.now()
    today = now.date()
    Product.objects.filter(
        current_stock__lte=F('reorder_level'), current_stock__gt=0, is_active=True
    ).exclude(
        stock_alerts__alert_type='low_stock', stock_alerts__status='active'
    ).values_list('id', flat=True)

    low_stock_products = Product.objects.filter(
        current_stock__lte=F('reorder_level'), current_stock__gt=0, is_active=True
    ).only('id', 'name', 'current_stock', 'reorder_level')

    for product in low_stock_products:
        StockAlert.objects.get_or_create(
            product=product,
            alert_type='low_stock',
            status='active',
            defaults={
                'message': f'{product.name} stock is low: {product.current_stock}',
                'current_value': product.current_stock,
                'threshold_value': product.reorder_level
            }
        )

    StockAlert.objects.filter(
        alert_type='low_stock', status='active'
    ).exclude(
        product__in=low_stock_products.values('id')
    ).update(status='resolved', resolved_at=now, resolved_by=request.user)

    expiring_products = Product.objects.filter(
        expiry_date__gte=today, expiry_date__lte=today + timedelta(days=90),
        is_active=True
    ).only('id', 'name', 'batch_number', 'expiry_date')

    for product in expiring_products:
        days_until = (product.expiry_date - today).days
        StockAlert.objects.get_or_create(
            product=product,
            alert_type='expiring_soon',
            status='active',
            defaults={
                'message': f'{product.name} batch {product.batch_number or "N/A"} will expire in {days_until} days',
                'current_value': days_until,
                'threshold_value': 90,
            }
        )

    expired_products = Product.objects.filter(
        expiry_date__lt=today, is_active=True
    ).only('id', 'name', 'batch_number', 'expiry_date')

    for product in expired_products:
        days_until = (product.expiry_date - today).days
        StockAlert.objects.get_or_create(
            product=product,
            alert_type='expired',
            status='active',
            defaults={
                'message': f'{product.name} batch {product.batch_number or "N/A"} has expired',
                'current_value': abs(days_until),
                'threshold_value': 0,
            }
        )

    active_ids = list(expiring_products.values_list('id', flat=True)) + list(expired_products.values_list('id', flat=True))
    StockAlert.objects.filter(
        alert_type__in=['expiring_soon', 'expired'], status='active'
    ).exclude(
        product__in=active_ids
    ).update(status='resolved', resolved_at=now, resolved_by=request.user)

    alerts = StockAlert.objects.select_related('product').filter(status='active')

    context = {
        'alerts': alerts,
        'alert_types': StockAlert.ALERT_TYPES,
        'total_alerts': alerts.count(),
    }
    return render(request, 'inventory/stock_alerts.html', context)


@login_required
def stock_adjustment_view(request):
    """Adjust stock levels manually"""
    if request.method == 'POST':
        form = StockAdjustmentForm(request.POST)
        if form.is_valid():
            adjustment = form.save(commit=False)
            adjustment.created_by = request.user

            product = adjustment.product
            previous_quantity = product.current_stock

            adjustment.previous_quantity = previous_quantity

            if adjustment.reason in ['damage', 'expiry', 'theft']:
                adjustment.new_quantity = previous_quantity - adjustment.quantity
            else:
                adjustment.new_quantity = previous_quantity + adjustment.quantity

            if adjustment.new_quantity < 0:
                messages.error(request, 'Cannot reduce stock below zero.')
                return render(request, 'inventory/stock_adjustment.html', {'form': form})

            adjustment.save()

            product.current_stock = adjustment.new_quantity
            product.save(update_fields=['current_stock'])

            StockMovement.objects.create(
                product=product,
                movement_type='adjustment',
                quantity=adjustment.quantity,
                previous_quantity=previous_quantity,
                new_quantity=adjustment.new_quantity,
                reference_type='Adjustment',
                reference_id=str(adjustment.id),
                notes=f"Reason: {adjustment.get_reason_display()}. {adjustment.notes}",
                created_by=request.user
            )

            if product.current_stock <= product.reorder_level and product.current_stock > 0:
                StockAlert.objects.create(
                    product=product,
                    alert_type='low_stock',
                    message=f'{product.name} stock is low: {product.current_stock} units remaining',
                    current_value=product.current_stock,
                    threshold_value=product.reorder_level
                )

            ActivityLog.objects.create(
                user=request.user,
                action='update',
                model_name='InventoryAdjustment',
                object_id=str(adjustment.id),
                object_repr=f"{product.name} adjustment",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, 'Stock adjusted successfully!')
            return redirect('inventory:stock_movement')
    else:
        form = StockAdjustmentForm()

    context = {'form': form}
    return render(request, 'inventory/stock_adjustment.html', context)


@login_required
def export_stock_pdf_view(request):
    """Generate PDF report of current stock levels based on active filters"""
    products = Product.objects.filter(is_active=True).select_related('category')

    search = request.GET.get('search', '')
    if search:
        products = products.filter(
            Q(name__icontains=search) |
            Q(generic_name__icontains=search) |
            Q(sku__icontains=search)
        )

    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)

    low_stock = request.GET.get('low_stock')
    if low_stock:
        products = products.filter(current_stock__lte=F('reorder_level'))

    context = {
        'products': products,
        'search': search,
        'generated_at': timezone.now(),
        'generated_by': request.user.get_full_name() or request.user.username,
        'total_items': products.count(),
    }

    template = get_template('inventory/stock_list_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    filename = f"Stock_List_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    return response


# ==============================================================================
# LOW STOCK REPORT VIEWS
# ==============================================================================

def _get_low_stock_products():
    """Helper to query all active low stock products"""
    return Product.objects.filter(
        current_stock__lte=F('reorder_level'),
        current_stock__gt=0,
        is_active=True
    ).select_related('category')


@login_required
def low_stock_report_view(request):
    """Generate low stock report page"""
    products = _get_low_stock_products()
    context = {'products': products}
    return render(request, 'inventory/low_stock_report.html', context)


@login_required
def export_low_stock_pdf_view(request):
    """Exports low stock report to PDF"""
    products = _get_low_stock_products()

    context = {
        'products': products,
        'generated_at': timezone.now(),
        'generated_by': request.user.get_full_name() or request.user.username,
        'total_items': products.count(),
    }

    template = get_template('inventory/low_stock_report_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    filename = f"Low_Stock_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    return response


@login_required
def export_low_stock_excel_view(request):
    """Exports low stock report to Excel (.xlsx)"""
    products = _get_low_stock_products()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Low Stock Report"

    header_fill = PatternFill(start_color="0D5C3A", end_color="0D5C3A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    headers = [
        "S/N",
        "Product Name",
        "Generic Name",
        "SKU",
        "Category",
        "Current Stock",
        "Reorder Level",
        "Shortage",
        "Supplier"
    ]

    worksheet.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, product in enumerate(products, start=1):
        category_name = product.category.name if product.category else "-"
        supplier_obj = getattr(product, 'supplier', None)
        supplier_name = supplier_obj.name if supplier_obj and hasattr(supplier_obj, 'name') else "-"
        shortage = getattr(product, 'reorder_quantity', None) or product.reorder_level

        row = [
            idx,
            product.name,
            product.generic_name or "-",
            product.sku,
            category_name,
            f"{product.current_stock} {product.unit}",
            f"{product.reorder_level} {product.unit}",
            f"{shortage} {product.unit}",
            supplier_name
        ]
        worksheet.append(row)

        row_num = idx + 1
        for col_num in range(1, len(headers) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [1, 4, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center")

    for col in worksheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Low_Stock_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response


# ==============================================================================
# CATEGORY & API VIEWS
# ==============================================================================

@login_required
def category_list_view(request):
    """List all categories"""
    categories = Category.objects.filter(parent__isnull=True)
    context = {'categories': categories}
    return render(request, 'inventory/category_list.html', context)


@login_required
def category_create_view(request):
    """Create a new category"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" created successfully!')
            return redirect('inventory:category_list')
    else:
        form = CategoryForm()

    context = {'form': form, 'title': 'Add Category'}
    return render(request, 'inventory/category_form.html', context)


@login_required
def barcode_lookup_view(request):
    """Look up product by barcode"""
    if request.method == 'GET' and request.GET.get('barcode'):
        barcode = request.GET.get('barcode')
        try:
            product = Product.objects.get(barcode=barcode, is_active=True)
            data = {
                'id': str(product.id),
                'name': product.name,
                'generic_name': product.generic_name,
                'sku': product.sku,
                'price': float(product.selling_price) if product.selling_price else 0,
                'stock': product.current_stock,
            }
            return JsonResponse(data)
        except Product.DoesNotExist:
            return JsonResponse({'error': 'Product not found'}, status=404)

    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def api_product_stock(request, product_id):
    """API endpoint to get product stock information"""
    try:
        product = get_object_or_404(Product, id=product_id)
        return JsonResponse({
            'id': str(product.id),
            'name': product.name,
            'current_stock': product.current_stock,
            'purchase_price': float(product.purchase_price),
            'selling_price': float(product.selling_price) if product.selling_price else 0,
            'markup_percent': float(product.discount_percent) if product.discount_percent else 0,
            'batch_number': product.batch_number or '',
            'expiry_date': product.expiry_date.isoformat() if product.expiry_date else '',
            'reorder_level': product.reorder_level,
            'is_low_stock': product.current_stock <= product.reorder_level,
            'is_out_of_stock': product.current_stock <= 0,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ==============================================================================
# SALES REPORT VIEWS & EXPORTS
# ==============================================================================

def _get_filtered_sales(request):
    """Helper function to filter sales items based on request parameters"""
    queryset = SaleItem.objects.select_related('sale', 'sale__customer', 'product').order_by('-sale__created_at')

    # Search filter (Product name or Customer name)
    search = request.GET.get('search', '').strip()
    if search:
        queryset = queryset.filter(
            Q(product__name__icontains=search) |
            Q(sale__customer__first_name__icontains=search) |
            Q(sale__customer__last_name__icontains=search) |
            Q(sale__customer_name__icontains=search)
        )

    # Payment Status filter
    payment_status = request.GET.get('payment_status', '').strip()
    if payment_status:
        queryset = queryset.filter(sale__payment_status__iexact=payment_status)

    # Date Range filters
    date_from = request.GET.get('date_from', '').strip()
    if date_from:
        queryset = queryset.filter(sale__created_at__date__gte=date_from)

    date_to = request.GET.get('date_to', '').strip()
    if date_to:
        queryset = queryset.filter(sale__created_at__date__lte=date_to)

    return queryset


def _get_item_amount_paid(item):
    """Helper to safely extract or calculate the amount paid for a line item"""
    if hasattr(item, 'total_price') and item.total_price is not None:
        return float(item.total_price)
    if hasattr(item, 'amount_paid') and item.amount_paid is not None:
        return float(item.amount_paid)
    if hasattr(item, 'unit_price') and item.unit_price is not None:
        return float(item.quantity * item.unit_price)
    if item.product and hasattr(item.product, 'selling_price') and item.product.selling_price:
        return float(item.quantity * item.product.selling_price)
    return 0.0


@login_required
def sales_report_view(request):
    """Renders the interactive Sales Report page with filters and pagination"""
    sales_items_qs = _get_filtered_sales(request)

    # Aggregate summaries across all matching filtered records
    total_records = sales_items_qs.count()
    totals = sales_items_qs.aggregate(
        total_quantity=Sum('quantity'),
        total_amount=Sum(F('unit_price') * F('quantity'))
    )
    total_quantity = totals['total_quantity'] or 0
    total_amount_paid = float(totals['total_amount'] or 0)

    paginator = Paginator(sales_items_qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Attach calculated amount paid for current page display
    for item in page_obj:
        item.calculated_amount_paid = _get_item_amount_paid(item)

    context = {
        'sales_items': page_obj,
        'search': request.GET.get('search', ''),
        'payment_status': request.GET.get('payment_status', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
        'total_records': total_records,
        'total_quantity': total_quantity,
        'total_amount_paid': total_amount_paid,
    }
    return render(request, 'inventory/sales_report.html', context)


@login_required
def export_sales_pdf_view(request):
    """Exports filtered sales report to PDF using xhtml2pdf"""
    sales_items_qs = _get_filtered_sales(request)

    sales_items_list = []
    total_quantity = 0
    total_amount_paid = 0.0

    for item in sales_items_qs:
        amt = _get_item_amount_paid(item)
        item.calculated_amount_paid = amt
        sales_items_list.append(item)
        total_quantity += item.quantity
        total_amount_paid += amt

    context = {
        'sales_items': sales_items_list,
        'generated_at': timezone.now(),
        'generated_by': request.user.get_full_name() or request.user.username,
        'total_records': len(sales_items_list),
        'total_quantity': total_quantity,
        'total_amount_paid': total_amount_paid,
        'search': request.GET.get('search', ''),
        'payment_status': request.GET.get('payment_status', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }

    template = get_template('inventory/sales_report_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    filename = f"Sales_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)

    return response


@login_required
def export_sales_excel_view(request):
    """Exports filtered sales report to Excel (.xlsx) using openpyxl"""
    sales_items_qs = _get_filtered_sales(request)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sales Report"
    worksheet.views.sheetView[0].showGridLines = True

    # Header & Row Styles
    header_fill = PatternFill(start_color="0D5C3A", end_color="0D5C3A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    total_fill = PatternFill(start_color="E6F0EC", end_color="E6F0EC", fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True, color="0D5C3A")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    total_border = Border(
        top=Side(style='thin', color='0D5C3A'),
        bottom=Side(style='double', color='0D5C3A'),
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB')
    )

    headers = [
        "S/N",
        "Date & Time",
        "Product Name",
        "Customer Name",
        "Quantity Sold",
        "Amount Paid",
        "Payment Status"
    ]

    worksheet.append(headers)

    # Style Headers
    for col_num in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add Data Rows
    start_row = 2
    for idx, item in enumerate(sales_items_qs, start=1):
        sale = item.sale

        # Resolve customer name
        if sale.customer:
            cust_name = sale.customer.get_full_name() if hasattr(sale.customer, 'get_full_name') and sale.customer.get_full_name() else str(sale.customer)
        else:
            cust_name = getattr(sale, 'customer_name', None) or 'Walk-in Customer'

        # Format creation date
        formatted_date = sale.created_at.strftime('%Y-%m-%d %H:%M') if sale.created_at else 'N/A'

        # Extract payment status & amount
        status = getattr(sale, 'payment_status', None) or getattr(sale, 'status', 'N/A')
        amount_paid = _get_item_amount_paid(item)

        row = [
            idx,
            formatted_date,
            item.product.name if item.product else 'Unknown Product',
            cust_name,
            item.quantity,
            amount_paid,
            str(status).capitalize()
        ]
        worksheet.append(row)

        row_num = start_row + idx - 1
        is_even = idx % 2 == 0

        for col_num in range(1, len(headers) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill

            if col_num in [1, 2, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_num == 6:
                    cell.number_format = 'TZS #,##0.00'

    # Summary Row
    last_data_row = start_row + len(sales_items_qs) - 1
    total_row_num = last_data_row + 1

    total_label_cell = worksheet.cell(row=total_row_num, column=1, value="TOTAL")
    total_label_cell.alignment = Alignment(horizontal="center", vertical="center")

    total_qty_cell = worksheet.cell(row=total_row_num, column=5, value=f"=SUM(E{start_row}:E{last_data_row})")
    total_qty_cell.alignment = Alignment(horizontal="right", vertical="center")

    total_amt_cell = worksheet.cell(row=total_row_num, column=6, value=f"=SUM(F{start_row}:F{last_data_row})")
    total_amt_cell.number_format = 'TZS #,##0.00'
    total_amt_cell.alignment = Alignment(horizontal="right", vertical="center")

    for col_num in range(1, len(headers) + 1):
        cell = worksheet.cell(row=total_row_num, column=col_num)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border

    # Auto-adjust Column Widths
    for col in worksheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Sales_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response